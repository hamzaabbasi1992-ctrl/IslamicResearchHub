import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

root_dir = Path(".")
docs_dir = Path("docs")

xlsx_files = list(root_dir.glob("*.xlsx")) + list(docs_dir.glob("*.xlsx"))

print(f"Scanning {len(xlsx_files)} Excel files for user highlights...")

for xfile in xlsx_files:
    try:
        wb = openpyxl.load_workbook(xfile)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            highlighted = []
            for r_idx, row in enumerate(ws.iter_rows(), start=1):
                for cell in row:
                    fill = cell.fill
                    if fill and fill.fill_type and fill.start_color:
                        sc = fill.start_color
                        color_val = str(sc.rgb or sc.theme or sc.indexed or "")
                        if color_val and color_val != "00000000" and color_val != "FFFFFFFF":
                            highlighted.append((r_idx, cell.coordinate, cell.value, color_val))

            if highlighted:
                print(f"\n📂 File: {xfile} | Sheet: {sheet_name} -> Found {len(highlighted)} styled/highlighted cells!")
                for h in highlighted[:10]:
                    print(f"   Cell {h[1]} (Row {h[0]}): Value='{h[2]}' | Color={h[3]}")
    except Exception as err:
        print(f"Error checking {xfile}: {err}")
