"""CLI & Engine to audit, report, and execute database deduplication.

Generates 3-section / 3-sheet audit files:
1. Deduplication_Audit_Total_Deleted_Remaining.xlsx
2. Deduplication_Audit_Total_Deleted_Remaining.csv
3. Deduplication_Audit_Total_Deleted_Remaining.txt

Then deletes redundant duplicate pages, chapters, and books from data/books.db and runs VACUUM to reclaim space.
"""

import argparse
import csv
import logging
import re
import sqlite3
from contextlib import closing
from pathlib import Path

import openpyxl

LOGGER = logging.getLogger(__name__)

DEFAULT_DATABASE_PATH = Path("data/books.db")
DEFAULT_EXCEL_AUDIT = Path("docs/Deduplication_Audit_Total_Deleted_Remaining.xlsx")
DEFAULT_CSV_AUDIT = Path("docs/Deduplication_Audit_Total_Deleted_Remaining.csv")
DEFAULT_TXT_AUDIT = Path("docs/Deduplication_Audit_Total_Deleted_Remaining.txt")


def normalize_title_key(title: str) -> str:
    """Normalize title by removing volume suffixes like (2), جلد 2, spaces, and punctuation."""
    t = str(title).strip().lower()
    t = re.sub(r"\s*\(\d+\)$", "", t)
    t = re.sub(r"\s*(جلد|المجلد|part|vol|volume)\s*\d+$", "", t)
    t = re.sub(r"[^\w\s\u0600-\u06FF]", "", t)
    return re.sub(r"\s+", " ", t).strip()


def plan_deduplication(conn: sqlite3.Connection) -> tuple[list[dict], list[dict], list[dict]]:
    """Group all books and return (total_books, remaining_books, deleted_books)."""
    conn.row_factory = sqlite3.Row
    books = conn.execute(
        """
        SELECT b.BookID,
               COALESCE(b.Title, 'Untitled') AS Title,
               COALESCE(b.Author, '') AS Author,
               COALESCE(l.Name, 'General') AS LibraryName,
               COALESCE(b.Category, 'General') AS Category,
               COALESCE(b.PageCount, 0) AS PageCount
        FROM Books b
        LEFT JOIN Libraries l ON l.LibraryID = b.LibraryID
        ORDER BY b.BookID ASC
        """
    ).fetchall()

    title_groups: dict[str, list[sqlite3.Row]] = {}
    for b in books:
        key = normalize_title_key(b["Title"])
        if not key:
            key = f"book_raw_{b['BookID']}"
        title_groups.setdefault(key, []).append(b)

    total_list = []
    remaining_list = []
    deleted_list = []

    for group_key, b_list in title_groups.items():
        # Sort by PageCount descending, then BookID ascending
        b_list.sort(key=lambda x: (x["PageCount"], -x["BookID"]), reverse=True)

        primary = b_list[0]
        rec_primary = {
            "BookID": primary["BookID"],
            "Title": primary["Title"],
            "Author": primary["Author"],
            "Maktaba": primary["LibraryName"],
            "Category": primary["Category"],
            "PageCount": primary["PageCount"],
            "ActionStatus": "REMAINING (Kept Primary)",
        }
        total_list.append(rec_primary)
        remaining_list.append(rec_primary)

        for redundant in b_list[1:]:
            rec_del = {
                "BookID": redundant["BookID"],
                "Title": redundant["Title"],
                "Author": redundant["Author"],
                "Maktaba": redundant["LibraryName"],
                "Category": redundant["Category"],
                "PageCount": redundant["PageCount"],
                "ActionStatus": f"DELETED (Duplicate of BookID #{primary['BookID']})",
            }
            total_list.append(rec_del)
            deleted_list.append(rec_del)

    # Sort total_list by BookID
    total_list.sort(key=lambda x: x["BookID"])
    remaining_list.sort(key=lambda x: x["BookID"])
    deleted_list.sort(key=lambda x: x["BookID"])

    return total_list, remaining_list, deleted_list


def generate_audit_files(
    total_list: list[dict],
    remaining_list: list[dict],
    deleted_list: list[dict],
    excel_path: Path = DEFAULT_EXCEL_AUDIT,
    csv_path: Path = DEFAULT_CSV_AUDIT,
    txt_path: Path = DEFAULT_TXT_AUDIT,
) -> None:
    """Generate Excel (.xlsx), CSV, and TXT audit files with Total, Remaining, and Deleted sections."""
    headers = ["BookID", "Title", "Author", "Maktaba", "Category", "PageCount", "ActionStatus"]

    # 1. Write TXT Audit File (Tab-Separated with Header Summaries)
    with open(txt_path, "w", encoding="utf-8") as f_txt:
        f_txt.write("=========================================================================\n")
        f_txt.write("ISLAMIC RESEARCH HUB AI - DEDUPLICATION MASTER AUDIT REPORT\n")
        f_txt.write("=========================================================================\n")
        f_txt.write(f"Total Books Cataloged:  {len(total_list):,}\n")
        f_txt.write(f"Remaining Books (Kept): {len(remaining_list):,}\n")
        f_txt.write(f"Deleted Books (Removed):{len(deleted_list):,}\n")
        f_txt.write("=========================================================================\n\n")

        f_txt.write("BookID\tTitle\tAuthor\tMaktaba\tCategory\tPageCount\tActionStatus\n")
        for rec in total_list:
            f_txt.write(f"{rec['BookID']}\t{rec['Title']}\t{rec['Author']}\t{rec['Maktaba']}\t{rec['Category']}\t{rec['PageCount']}\t{rec['ActionStatus']}\n")

    # 2. Write CSV Audit File
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f_csv:
        writer = csv.writer(f_csv)
        writer.writerow(headers)
        for rec in total_list:
            writer.writerow([
                rec["BookID"], rec["Title"], rec["Author"], rec["Maktaba"],
                rec["Category"], rec["PageCount"], rec["ActionStatus"]
            ])

    # 3. Write Excel (.xlsx) Audit File with 3 Separate Sheets
    wb = openpyxl.Workbook(write_only=True)

    ws_total = wb.create_sheet(title=f"Total Books ({len(total_list):,})")
    ws_total.append(headers)
    for rec in total_list:
        ws_total.append([rec["BookID"], rec["Title"], rec["Author"], rec["Maktaba"], rec["Category"], rec["PageCount"], rec["ActionStatus"]])

    ws_rem = wb.create_sheet(title=f"Remaining ({len(remaining_list):,})")
    ws_rem.append(headers)
    for rec in remaining_list:
        ws_rem.append([rec["BookID"], rec["Title"], rec["Author"], rec["Maktaba"], rec["Category"], rec["PageCount"], rec["ActionStatus"]])

    ws_del = wb.create_sheet(title=f"Deleted ({len(deleted_list):,})")
    ws_del.append(headers)
    for rec in deleted_list:
        ws_del.append([rec["BookID"], rec["Title"], rec["Author"], rec["Maktaba"], rec["Category"], rec["PageCount"], rec["ActionStatus"]])

    try:
        wb.save(excel_path)
    except PermissionError:
        excel_path = excel_path.parent / f"{excel_path.stem}_Updated{excel_path.suffix}"
        wb.save(excel_path)


def execute_deduplication(database_path: Path, deleted_list: list[dict]) -> tuple[int, int]:
    """Delete redundant books, pages, and chapters from database and VACUUM."""
    deleted_ids = [rec["BookID"] for rec in deleted_list]
    if not deleted_ids:
        return 0, 0

    with closing(sqlite3.connect(database_path)) as conn:
        conn.execute("PRAGMA foreign_keys = OFF;")

        # Delete in chunks to avoid SQLite parameter limit
        chunk_size = 500
        total_pages_deleted = 0
        total_books_deleted = 0

        for i in range(0, len(deleted_ids), chunk_size):
            chunk = deleted_ids[i : i + chunk_size]
            placeholders = ",".join("?" * len(chunk))

            cur_p = conn.execute(f"DELETE FROM Pages WHERE BookID IN ({placeholders})", chunk)
            total_pages_deleted += cur_p.rowcount

            conn.execute(f"DELETE FROM Chapters WHERE BookID IN ({placeholders})", chunk)

            cur_b = conn.execute(f"DELETE FROM Books WHERE BookID IN ({placeholders})", chunk)
            total_books_deleted += cur_b.rowcount

            conn.commit()

        # Attempt VACUUM if sufficient temp disk space is available
        try:
            print("Reclaiming disk space via VACUUM...")
            conn.execute("VACUUM;")
        except sqlite3.OperationalError:
            freelist = conn.execute("PRAGMA freelist_count;").fetchone()[0]
            page_size = conn.execute("PRAGMA page_size;").fetchone()[0]
            freelist_gb = (freelist * page_size) / (1024 * 1024 * 1024)
            LOGGER.info("VACUUM skipped due to temp disk space limit. Freelist available inside DB: %.2f GB (%d pages)", freelist_gb, freelist)

    return total_books_deleted, total_pages_deleted


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit, export reports, and execute database deduplication")
    parser.add_argument("--database", type=Path, default=DEFAULT_DATABASE_PATH)
    parser.add_argument("--excel-audit", type=Path, default=DEFAULT_EXCEL_AUDIT)
    parser.add_argument("--csv-audit", type=Path, default=DEFAULT_CSV_AUDIT)
    parser.add_argument("--txt-audit", type=Path, default=DEFAULT_TXT_AUDIT)
    parser.add_argument("--execute", action="store_true", help="Execute actual deletion and VACUUM")
    args = parser.parse_args()

    if not args.database.is_file():
        print(f"Error: Database file not found: {args.database}")
        return

    with closing(sqlite3.connect(args.database)) as conn:
        total_list, remaining_list, deleted_list = plan_deduplication(conn)

    print("Deduplication Plan Calculated:")
    print(f"Total Books:           {len(total_list):,}")
    print(f"Remaining Books (Kept): {len(remaining_list):,}")
    print(f"Deleted Books (Remove): {len(deleted_list):,}")

    print("\nGenerating Audit Reports...")
    generate_audit_files(total_list, remaining_list, deleted_list, args.excel_audit, args.csv_audit, args.txt_audit)
    print(f"1. TXT Report:   {args.txt_audit.resolve()}")
    print(f"2. CSV Report:   {args.csv_audit.resolve()}")
    print(f"3. Excel Report: {args.excel_audit.resolve()}")

    if args.execute:
        print("\nExecuting Database Deduplication & VACUUM...")
        books_del, pages_del = execute_deduplication(args.database, deleted_list)
        print(f"\nDeduplication Completed Successfully!")
        print(f"Deleted Books: {books_del:,}")
        print(f"Deleted Pages: {pages_del:,}")
    else:
        print("\n[Notice] Run with --execute flag to perform actual database deletion and space recovery.")


if __name__ == "__main__":
    main()
