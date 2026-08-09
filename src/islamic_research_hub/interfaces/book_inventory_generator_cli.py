"""CLI & Utility to update all CSV and Excel (.xlsx) inventory files in docs/book_inventory.

Generates both editable CSV and Excel (.xlsx) formats for:
1. docs/book_inventory/all_books.csv / .xlsx
2. docs/book_inventory/searchable_books_arabic.csv / .xlsx
3. docs/book_inventory/searchable_books_urdu.csv / .xlsx
4. docs/book_inventory/searchable_books_english.csv / .xlsx
5. docs/book_inventory/searchable_books_unspecified.csv / .xlsx
6. docs/book_inventory/multi_volume_series.csv / .xlsx
7. docs/book_inventory/missing_volumes_availability.csv / .xlsx
"""

import argparse
import csv
import logging
import sqlite3
from contextlib import closing
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

LOGGER = logging.getLogger(__name__)

DEFAULT_DATABASE_PATH = Path("data/books.db")
DEFAULT_INVENTORY_DIR = Path("docs/book_inventory")


def _write_csv_and_xlsx(
    csv_path: Path,
    xlsx_path: Path,
    headers: list[str],
    data_rows: list[list[object]],
    sheet_title: str = "Inventory",
) -> None:
    """Write data to UTF-8-sig CSV and styled editable Excel (.xlsx) file."""
    # Write CSV
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data_rows)

    # Fast Write Excel (.xlsx)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_title)

    ws.append(headers)
    for row in data_rows:
        ws.append(row)

    wb.save(xlsx_path)


def update_book_inventory_files(
    database_path: Path = DEFAULT_DATABASE_PATH,
    inventory_dir: Path = DEFAULT_INVENTORY_DIR,
) -> dict[str, int]:
    """Generate and overwrite all CSV and Excel inventory files in docs/book_inventory."""
    if not database_path.is_file():
        raise FileNotFoundError(f"Database file not found: {database_path}")

    inventory_dir.mkdir(parents=True, exist_ok=True)

    with closing(sqlite3.connect(database_path)) as conn:
        conn.row_factory = sqlite3.Row
        books = conn.execute(
            """
            SELECT b.BookID,
                   COALESCE(b.Title, 'Untitled') AS Title,
                   COALESCE(b.Author, '') AS Author,
                   COALESCE(l.Name, 'General') AS LibraryName,
                   COALESCE(b.Language, '') AS Language,
                   COALESCE(b.PageCount, 0) AS PageCount,
                   b.SeriesID,
                   b.VolumeNumber
            FROM Books b
            LEFT JOIN Libraries l ON l.LibraryID = b.LibraryID
            ORDER BY b.BookID ASC
            """
        ).fetchall()

    counts = {
        "all": len(books),
        "ar": 0,
        "ur": 0,
        "en": 0,
        "unspecified": 0,
        "series": 0,
    }

    # Data structures for each file
    all_rows = []
    ar_rows = []
    ur_rows = []
    en_rows = []
    unspecified_rows = []

    std_headers = ["BookID", "Title", "Author", "Maktaba", "Language", "PageCount", "ContentStatus"]

    for row in books:
        status = "searchable text" if row["PageCount"] > 0 else "metadata only"
        record = [
            row["BookID"],
            row["Title"],
            row["Author"],
            row["LibraryName"],
            row["Language"],
            row["PageCount"],
            status,
        ]
        all_rows.append(record)

        lang = (row["Language"] or "").strip().lower()
        if lang in ("ar", "arabic"):
            ar_rows.append(record)
            counts["ar"] += 1
        elif lang in ("ur", "urdu"):
            ur_rows.append(record)
            counts["ur"] += 1
        elif lang in ("en", "english"):
            en_rows.append(record)
            counts["en"] += 1
        else:
            unspecified_rows.append(record)
            counts["unspecified"] += 1

    # 1. Save all_books
    _write_csv_and_xlsx(
        inventory_dir / "all_books.csv",
        inventory_dir / "all_books.xlsx",
        std_headers,
        all_rows,
        "All Books",
    )

    # 2. Save language specific inventories
    _write_csv_and_xlsx(
        inventory_dir / "searchable_books_arabic.csv",
        inventory_dir / "searchable_books_arabic.xlsx",
        std_headers,
        ar_rows,
        "Arabic Books",
    )
    _write_csv_and_xlsx(
        inventory_dir / "searchable_books_urdu.csv",
        inventory_dir / "searchable_books_urdu.xlsx",
        std_headers,
        ur_rows,
        "Urdu Books",
    )
    _write_csv_and_xlsx(
        inventory_dir / "searchable_books_english.csv",
        inventory_dir / "searchable_books_english.xlsx",
        std_headers,
        en_rows,
        "English Books",
    )
    _write_csv_and_xlsx(
        inventory_dir / "searchable_books_unspecified.csv",
        inventory_dir / "searchable_books_unspecified.xlsx",
        std_headers,
        unspecified_rows,
        "Unspecified Books",
    )

    # 3. Save Series inventory files
    series_groups: dict[int, list[sqlite3.Row]] = {}
    for row in books:
        sid = row["SeriesID"]
        if sid is not None and sid > 0:
            series_groups.setdefault(sid, []).append(row)

    counts["series"] = len(series_groups)

    series_headers = [
        "SeriesID", "Title", "Author", "Maktaba", "VolumesHave",
        "FirstVol", "LastVol", "MissingVolumeNumbers", "Status",
        "SourceFileCount", "Confidence"
    ]
    missing_headers = [
        "SeriesID", "Title", "Author", "Maktaba",
        "MissingVolumeNumbers", "Availability", "Notes", "SourceURL"
    ]

    series_rows = []
    missing_rows = []

    for sid, s_books in sorted(series_groups.items()):
        first_b = s_books[0]
        title = first_b["Title"]
        author = first_b["Author"]
        maktaba = first_b["LibraryName"]

        vols = [b["VolumeNumber"] for b in s_books if b["VolumeNumber"] is not None]
        if not vols:
            continue
        vols = sorted(set(vols))
        min_v, max_v = min(vols), max(vols)
        expected = set(range(min_v, max_v + 1))
        missing = sorted(expected - set(vols))

        missing_str = ",".join(str(m) for m in missing) if missing else ""
        status = "COMPLETE" if not missing else "INCOMPLETE"

        series_rows.append([
            sid, title, author, maktaba, len(vols),
            min_v, max_v, missing_str, status, len(s_books),
            "HIGH - verified database series"
        ])

        if missing:
            missing_rows.append([
                sid, title, author, maktaba,
                missing_str, "Not researched", "Gaps detected in volume sequence", ""
            ])

    _write_csv_and_xlsx(
        inventory_dir / "multi_volume_series.csv",
        inventory_dir / "multi_volume_series.xlsx",
        series_headers,
        series_rows,
        "Multi-Volume Series",
    )
    _write_csv_and_xlsx(
        inventory_dir / "missing_volumes_availability.csv",
        inventory_dir / "missing_volumes_availability.xlsx",
        missing_headers,
        missing_rows,
        "Missing Volumes",
    )

    return counts


def main() -> None:
    parser = argparse.ArgumentParser(description="Update all CSV and Excel (.xlsx) inventory files in docs/book_inventory")
    parser.add_argument("--database", type=Path, default=DEFAULT_DATABASE_PATH)
    parser.add_argument("--inventory-dir", type=Path, default=DEFAULT_INVENTORY_DIR)
    args = parser.parse_args()

    counts = update_book_inventory_files(args.database, args.inventory_dir)
    print(f"Book Inventory Update Complete:")
    print(f"Total Books: {counts['all']:,}")
    print(f"Arabic: {counts['ar']:,}")
    print(f"Urdu: {counts['ur']:,}")
    print(f"English: {counts['en']:,}")
    print(f"Unspecified: {counts['unspecified']:,}")
    print(f"Series Tracked: {counts['series']:,}")


if __name__ == "__main__":
    main()
